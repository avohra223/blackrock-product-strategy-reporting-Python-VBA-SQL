"""
generate_factsheet.py
Generates a formatted Excel factsheet from attribution outputs.

Produces a multi-tab Excel workbook with:
  - Summary: Key metrics, cumulative performance, attribution breakdown
  - Monthly Detail: Month-by-month attribution by sector
  - AUM & Flows: Asset tracking with visual charts
  - Sector Attribution: Heatmap-style sector contribution analysis

Designed to simulate the kind of client-ready collateral a Product Strategist
would produce for portfolio managers and institutional clients.
"""

import pandas as pd
import numpy as np
import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# STYLING CONSTANTS
# ─────────────────────────────────────────────

# BlackRock-inspired colour palette
BLACK = "000000"
DARK_GREY = "333333"
MED_GREY = "666666"
LIGHT_GREY = "F2F2F2"
ACCENT_GREEN = "00875A"
ACCENT_RED = "D32F2F"
HEADER_BG = "1A1A2E"
HEADER_FG = "FFFFFF"
SUBHEADER_BG = "E8E8E8"
BORDER_COLOR = "CCCCCC"

HEADER_FONT = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10, color=DARK_GREY)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=BLACK)
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=10, color=MED_GREY)
DATA_FONT = Font(name="Calibri", size=10, color=DARK_GREY)
METRIC_FONT = Font(name="Calibri", bold=True, size=11, color=BLACK)
POSITIVE_FONT = Font(name="Calibri", bold=True, size=10, color=ACCENT_GREEN)
NEGATIVE_FONT = Font(name="Calibri", bold=True, size=10, color=ACCENT_RED)

HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR)
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

def apply_header_row(ws, row, cols, headers):
    """Apply formatted headers to a row."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=cols[i], value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def apply_data_cell(ws, row, col, value, fmt=None, font=None):
    """Write a formatted data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or DATA_FONT
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def write_metric_pair(ws, row, col, label, value, fmt="0.00%"):
    """Write a label-value metric pair."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = SUBHEADER_FONT
    label_cell.alignment = LEFT
    label_cell.border = THIN_BORDER
    label_cell.fill = LIGHT_FILL

    val_cell = ws.cell(row=row, column=col + 1, value=value)
    val_cell.font = METRIC_FONT
    val_cell.alignment = RIGHT
    val_cell.border = THIN_BORDER
    if fmt:
        val_cell.number_format = fmt


# ─────────────────────────────────────────────
# SHEET 1: SUMMARY
# ─────────────────────────────────────────────

def build_summary_sheet(wb, metadata, risk_metrics, monthly_summary, cumulative):
    """Build the executive summary sheet."""
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1A1A2E"

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    # Title
    ws.cell(row=2, column=2, value=metadata["fund_name"]).font = TITLE_FONT
    ws.cell(row=3, column=2, value=f"Performance Report | {metadata['reporting_period']}").font = SUBTITLE_FONT
    ws.cell(row=4, column=2, value=f"Benchmark: {metadata['benchmark']}").font = SUBTITLE_FONT

    # Key Metrics section
    row = 6
    ws.cell(row=row, column=2, value="KEY METRICS").font = Font(name="Calibri", bold=True, size=11, color=HEADER_BG)
    row += 1

    metrics_layout = [
        ("Cumulative Fund Return", risk_metrics["cumulative_fund_return"] / 100),
        ("Cumulative Benchmark Return", risk_metrics["cumulative_benchmark_return"] / 100),
        ("Cumulative Active Return", risk_metrics["cumulative_active_return"] / 100),
        ("Annualized Fund Return", risk_metrics["annualized_fund_return"] / 100),
        ("Annualized Volatility (Fund)", risk_metrics["annualized_fund_volatility"] / 100),
        ("Tracking Error", risk_metrics["tracking_error"] / 100),
        ("Information Ratio", risk_metrics["information_ratio"]),
        ("Sharpe Ratio", risk_metrics["sharpe_ratio"]),
        ("Max Drawdown", risk_metrics["max_drawdown"] / 100),
        ("% Positive Months", risk_metrics["pct_positive_months"] / 100),
    ]

    for label, value in metrics_layout:
        fmt = "0.00x" if "Ratio" in label else "0.00%"
        if "Ratio" in label:
            fmt = "0.00"
        write_metric_pair(ws, row, 2, label, value, fmt)
        row += 1

    # Monthly returns table
    row += 2
    ws.cell(row=row, column=2, value="MONTHLY RETURNS").font = Font(name="Calibri", bold=True, size=11, color=HEADER_BG)
    row += 1

    headers = ["Month", "Fund Return", "Benchmark Return", "Active Return"]
    apply_header_row(ws, row, [2, 3, 4, 5], headers)
    row += 1

    for _, mrow in monthly_summary.iterrows():
        ws.cell(row=row, column=2, value=mrow["date"].strftime("%b %Y")).font = DATA_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).border = THIN_BORDER

        apply_data_cell(ws, row, 3, mrow["fund_return"], "0.00%")
        apply_data_cell(ws, row, 4, mrow["benchmark_return"], "0.00%")

        active = mrow["active_return"]
        font = POSITIVE_FONT if active >= 0 else NEGATIVE_FONT
        apply_data_cell(ws, row, 5, active, "0.00%", font)
        row += 1

    # Cumulative performance chart
    chart_row = row + 2
    ws.cell(row=chart_row, column=2, value="CUMULATIVE PERFORMANCE").font = Font(name="Calibri", bold=True, size=11, color=HEADER_BG)

    # Write chart data
    data_start = chart_row + 1
    ws.cell(row=data_start, column=2, value="Month").font = HEADER_FONT
    ws.cell(row=data_start, column=2).fill = HEADER_FILL
    ws.cell(row=data_start, column=3, value="Fund").font = HEADER_FONT
    ws.cell(row=data_start, column=3).fill = HEADER_FILL
    ws.cell(row=data_start, column=4, value="Benchmark").font = HEADER_FONT
    ws.cell(row=data_start, column=4).fill = HEADER_FILL

    for i, (_, crow) in enumerate(cumulative.iterrows()):
        r = data_start + 1 + i
        ws.cell(row=r, column=2, value=crow["date"].strftime("%b %Y"))
        ws.cell(row=r, column=3, value=crow["cum_fund_return"])
        ws.cell(row=r, column=3).number_format = "0.00%"
        ws.cell(row=r, column=4, value=crow["cum_benchmark_return"])
        ws.cell(row=r, column=4).number_format = "0.00%"

    chart = LineChart()
    chart.title = "Cumulative Returns: Fund vs Benchmark"
    chart.style = 10
    chart.y_axis.title = "Cumulative Return"
    chart.y_axis.numFmt = "0.0%"
    chart.x_axis.title = "Month"
    chart.width = 20
    chart.height = 12

    cats = Reference(ws, min_col=2, min_row=data_start + 1, max_row=data_start + len(cumulative))
    fund_data = Reference(ws, min_col=3, min_row=data_start, max_row=data_start + len(cumulative))
    bench_data = Reference(ws, min_col=4, min_row=data_start, max_row=data_start + len(cumulative))

    chart.add_data(fund_data, titles_from_data=True)
    chart.add_data(bench_data, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = "1A1A2E"
    chart.series[1].graphicalProperties.line.solidFill = "999999"
    chart.series[1].graphicalProperties.line.dashStyle = "dash"

    ws.add_chart(chart, f"F7")


# ─────────────────────────────────────────────
# SHEET 2: SECTOR ATTRIBUTION
# ─────────────────────────────────────────────

def build_attribution_sheet(wb, attribution_df):
    """Build sector-level attribution detail sheet."""
    ws = wb.create_sheet("Sector Attribution")
    ws.sheet_properties.tabColor = "00875A"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14

    ws.cell(row=2, column=2, value="Sector Attribution Analysis").font = TITLE_FONT
    ws.cell(row=3, column=2, value="Brinson-Hood-Beebower Decomposition (Full Period Aggregate)").font = SUBTITLE_FONT

    # Aggregate by sector
    sector_agg = attribution_df.groupby("sector").agg({
        "fund_weight": "mean",
        "benchmark_weight": "mean",
        "active_weight": "mean",
        "allocation_effect": "sum",
        "selection_effect": "sum",
        "interaction_effect": "sum",
        "total_effect": "sum"
    }).sort_values("total_effect", ascending=False)

    row = 5
    headers = ["Sector", "Avg Fund Wt", "Avg BM Wt", "Active Wt",
               "Allocation", "Selection", "Interaction", "Total Effect"]
    apply_header_row(ws, row, list(range(2, 10)), headers)
    row += 1

    for sector, srow in sector_agg.iterrows():
        ws.cell(row=row, column=2, value=sector).font = DATA_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).border = THIN_BORDER

        apply_data_cell(ws, row, 3, srow["fund_weight"], "0.0%")
        apply_data_cell(ws, row, 4, srow["benchmark_weight"], "0.0%")

        aw = srow["active_weight"]
        aw_font = POSITIVE_FONT if aw >= 0 else NEGATIVE_FONT
        apply_data_cell(ws, row, 5, aw, "+0.0%;-0.0%", aw_font)

        for j, col_name in enumerate(["allocation_effect", "selection_effect", "interaction_effect", "total_effect"]):
            val = srow[col_name]
            font = POSITIVE_FONT if val >= 0 else NEGATIVE_FONT
            apply_data_cell(ws, row, 6 + j, val, "0.00%", font)

            # Conditional fill
            if val >= 0.005:
                ws.cell(row=row, column=6 + j).fill = GREEN_FILL
            elif val <= -0.005:
                ws.cell(row=row, column=6 + j).fill = RED_FILL

        row += 1

    # Totals row
    ws.cell(row=row, column=2, value="TOTAL").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row, column=2).fill = SUBHEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER

    for j, col_name in enumerate(["allocation_effect", "selection_effect", "interaction_effect", "total_effect"]):
        total = sector_agg[col_name].sum()
        cell = apply_data_cell(ws, row, 6 + j, total, "0.00%",
                              Font(name="Calibri", bold=True, size=10, color=BLACK))
        cell.fill = SUBHEADER_FILL

    for col in range(3, 6):
        ws.cell(row=row, column=col).fill = SUBHEADER_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER

    # Attribution bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Sector Total Attribution Effect"
    chart.y_axis.title = "Contribution"
    chart.y_axis.numFmt = "0.00%"
    chart.width = 22
    chart.height = 14
    chart.style = 10

    # Chart data -- total effect column
    data_ref = Reference(ws, min_col=9, min_row=5, max_row=5 + len(sector_agg))
    cats_ref = Reference(ws, min_col=2, min_row=6, max_row=5 + len(sector_agg))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "1A1A2E"

    ws.add_chart(chart, "B" + str(row + 3))


# ─────────────────────────────────────────────
# SHEET 3: MONTHLY DETAIL
# ─────────────────────────────────────────────

def build_monthly_detail_sheet(wb, attribution_df):
    """Build month-by-month sector attribution detail."""
    ws = wb.create_sheet("Monthly Detail")
    ws.sheet_properties.tabColor = "FF8F00"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14

    ws.cell(row=2, column=2, value="Monthly Sector Attribution Detail").font = TITLE_FONT
    ws.cell(row=3, column=2, value="Total Effect by Sector per Month").font = SUBTITLE_FONT

    # Pivot: months as rows, sectors as columns
    pivot = attribution_df.pivot_table(
        index="date", columns="sector", values="total_effect", aggfunc="sum"
    )

    row = 5
    # Headers
    ws.cell(row=row, column=2, value="Month").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER

    for j, sector in enumerate(pivot.columns):
        col = 3 + j
        ws.column_dimensions[get_column_letter(col)].width = 14
        cell = ws.cell(row=row, column=col, value=sector)
        cell.font = Font(name="Calibri", bold=True, size=8, color=HEADER_FG)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row += 1

    for date, prow in pivot.iterrows():
        ws.cell(row=row, column=2, value=date.strftime("%b %Y")).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        for j, sector in enumerate(pivot.columns):
            val = prow[sector]
            font = POSITIVE_FONT if val >= 0 else NEGATIVE_FONT
            apply_data_cell(ws, row, 3 + j, val, "0.00%", font)

            if val >= 0.002:
                ws.cell(row=row, column=3 + j).fill = GREEN_FILL
            elif val <= -0.002:
                ws.cell(row=row, column=3 + j).fill = RED_FILL

        row += 1


# ─────────────────────────────────────────────
# SHEET 4: AUM & FLOWS
# ─────────────────────────────────────────────

def build_aum_sheet(wb, aum_df, metadata):
    """Build AUM and flows tracking sheet."""
    ws = wb.create_sheet("AUM & Flows")
    ws.sheet_properties.tabColor = "1565C0"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    ws.cell(row=2, column=2, value=f"{metadata['fund_name']}").font = TITLE_FONT
    ws.cell(row=3, column=2, value="Assets Under Management & Net Flows ($M)").font = SUBTITLE_FONT

    row = 5
    headers = ["Month", "AUM Start", "Net Flows", "Market Return", "AUM End"]
    apply_header_row(ws, row, [2, 3, 4, 5, 6], headers)
    row += 1

    for _, arow in aum_df.iterrows():
        ws.cell(row=row, column=2, value=arow["date"].strftime("%b %Y")).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        apply_data_cell(ws, row, 3, arow["aum_start_mm"], "#,##0.0")
        
        flow = arow["net_flows_mm"]
        flow_font = POSITIVE_FONT if flow >= 0 else NEGATIVE_FONT
        apply_data_cell(ws, row, 4, flow, "+#,##0.0;-#,##0.0", flow_font)

        apply_data_cell(ws, row, 5, arow["market_return_pct"] / 100, "0.00%")
        apply_data_cell(ws, row, 6, arow["aum_end_mm"], "#,##0.0")
        row += 1

    # AUM line chart
    chart = LineChart()
    chart.title = "AUM Trend ($M)"
    chart.style = 10
    chart.y_axis.title = "AUM ($M)"
    chart.y_axis.numFmt = "#,##0"
    chart.width = 18
    chart.height = 12

    cats = Reference(ws, min_col=2, min_row=6, max_row=5 + len(aum_df))
    aum_data = Reference(ws, min_col=6, min_row=5, max_row=5 + len(aum_df))
    chart.add_data(aum_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = "1565C0"

    ws.add_chart(chart, "B" + str(row + 2))


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    output_dir = os.path.join(os.path.dirname(__file__), "..", "output")
    data_dir = os.path.join(os.path.dirname(__file__), "..", "data")

    # Load outputs from attribution engine
    attribution_df = pd.read_csv(
        os.path.join(output_dir, "sector_attribution.csv"),
        parse_dates=["date"]
    )
    monthly_summary = pd.read_csv(
        os.path.join(output_dir, "monthly_summary.csv"),
        parse_dates=["date"]
    )
    cumulative = pd.read_csv(
        os.path.join(output_dir, "cumulative_performance.csv"),
        parse_dates=["date"]
    )
    with open(os.path.join(output_dir, "risk_metrics.json"), "r") as f:
        risk_metrics = json.load(f)

    aum_df = pd.read_csv(
        os.path.join(data_dir, "aum_flows.csv"),
        parse_dates=["date"]
    )
    with open(os.path.join(data_dir, "fund_metadata.json"), "r") as f:
        metadata = json.load(f)

    # Build workbook
    print("Building Excel factsheet...")
    wb = Workbook()

    build_summary_sheet(wb, metadata, risk_metrics, monthly_summary, cumulative)
    build_attribution_sheet(wb, attribution_df)
    build_monthly_detail_sheet(wb, attribution_df)
    build_aum_sheet(wb, aum_df, metadata)

    output_path = os.path.join(output_dir, "Fund_Factsheet_Report.xlsx")
    wb.save(output_path)
    print(f"Factsheet saved to {output_path}")
